from __future__ import annotations

import time
from collections import deque
from pathlib import Path
from typing import Any

from ..stats import ColumnAccumulator
from ..summarizer import (
    ColumnInfo,
    ColumnStats,
    DataSummary,
    SummarizerOptions,
    TableSummary,
)
from .base import MissingOptionalDep, Reader


class ExcelReader(Reader):
    FORMAT_NAME = "xlsx"

    def read(self, path: Path, opts: SummarizerOptions) -> DataSummary:
        try:
            from openpyxl import load_workbook  # noqa: WPS433
        except ImportError:
            raise MissingOptionalDep(
                "Excel support requires openpyxl. Install with: pip install openpyxl"
            )

        t0 = time.monotonic()
        size = path.stat().st_size
        warnings: list[str] = []

        wb = load_workbook(filename=str(path), read_only=True, data_only=True)
        all_sheets = wb.sheetnames

        if opts.tables:
            selected = [s for s in all_sheets if s in opts.tables]
            missing = [s for s in opts.tables if s not in all_sheets]
            if missing:
                warnings.append(f"Sheets not found: {', '.join(missing)}")
        else:
            selected = all_sheets

        if not opts.all_tables and len(selected) > opts.max_tables:
            warnings.append(
                f"Showing first {opts.max_tables} of {len(selected)} sheets "
                f"(use --all-tables or --max-tables to change)"
            )
            selected = selected[: opts.max_tables]

        tables: list[TableSummary] = []
        for sheet_name in selected:
            ws = wb[sheet_name]
            tables.append(self._summarize_sheet(sheet_name, ws, opts))

        wb.close()

        return DataSummary(
            source=str(path),
            file_size_bytes=size,
            file_format="xlsx",
            parse_duration_ms=int((time.monotonic() - t0) * 1000),
            backend_used="openpyxl",
            tables=tables,
            warnings=warnings,
        )

    def _summarize_sheet(self, name: str, ws: Any, opts: SummarizerOptions) -> TableSummary:
        rows_iter = ws.iter_rows(values_only=True)
        try:
            header_row = next(rows_iter)
        except StopIteration:
            return TableSummary(
                name=name,
                row_count=0,
                column_count=0,
                columns=[],
                stats=[],
                head=[],
                tail=[],
                notes=["Empty sheet"],
            )

        header = [str(h) if h is not None else f"col_{i}" for i, h in enumerate(header_row)]
        if opts.columns:
            col_indices = [(i, h) for i, h in enumerate(header) if h in opts.columns]
        else:
            col_indices = list(enumerate(header))
        col_indices = col_indices[: opts.max_columns]
        column_names = [h for _, h in col_indices]

        accumulators: dict[str, ColumnAccumulator] = {
            n: ColumnAccumulator(
                name=n,
                max_distinct=opts.max_distinct,
                keep_samples_for_median=opts.median,
            )
            for n in column_names
        }

        head_rows: list[dict[str, Any]] = []
        tail_maxlen = opts.sample_tail if (not opts.no_sample and opts.sample_tail > 0) else 0
        tail_buf: deque[dict[str, Any]] = deque(maxlen=tail_maxlen) if tail_maxlen else deque()

        row_count = 0
        truncated = False
        for row in rows_iter:
            row_count += 1
            if row_count > opts.max_rows:
                truncated = True
                break

            row_dict: dict[str, Any] = {}
            for i, n in col_indices:
                val = row[i] if i < len(row) else None
                row_dict[n] = val
                accumulators[n].update(val)

            if not opts.no_sample:
                if len(head_rows) < opts.sample_head:
                    head_rows.append(row_dict)
                if tail_maxlen:
                    tail_buf.append(row_dict)

        columns: list[ColumnInfo] = []
        stats: list[ColumnStats] = []
        for n in column_names:
            acc = accumulators[n]
            columns.append(
                ColumnInfo(
                    name=n,
                    dtype=acc.dtype(),
                    nullable=acc.null_count > 0,
                    null_count=acc.null_count,
                    null_pct=acc.null_pct(),
                    distinct_count=acc.distinct_count,
                )
            )
            if not opts.no_stats:
                top_vals = (
                    acc.top_values(opts.top_k)
                    if acc.dtype() not in ("int", "float", "datetime", "null")
                    else []
                )
                stats.append(
                    ColumnStats(
                        name=n,
                        count=acc.count,
                        null_count=acc.null_count,
                        distinct_count=acc.distinct_count,
                        min=acc.min_val,
                        max=acc.max_val,
                        mean=acc.mean,
                        median=acc.median if opts.median else None,
                        std=acc.std,
                        min_date=acc.min_dt,
                        max_date=acc.max_dt,
                        top_values=top_vals,
                    )
                )

        tail_rows = list(tail_buf)
        if head_rows and tail_rows and row_count <= opts.sample_head + len(tail_rows):
            tail_rows = []

        notes: list[str] = []
        if truncated:
            notes.append(f"Stats sampled from first {opts.max_rows:,} rows")

        return TableSummary(
            name=name,
            row_count=None if truncated else row_count,
            column_count=len(column_names),
            columns=columns,
            stats=stats,
            head=head_rows,
            tail=tail_rows,
            truncated=truncated,
            notes=notes,
        )
